"""
Excel文件解析器
Excel File Parser
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import logging
from typing import Dict, List, Any, Optional, Tuple
import os
import json
from datetime import datetime

logger = logging.getLogger(__name__)

class ExcelParser:
    """Excel文件解析器"""

    def __init__(self):
        self.config_rules = {}
        self.load_default_rules()

    def load_default_rules(self):
        """加载默认解析规则"""
        self.config_rules = {
            'vehicle_info': {
                'sheet_patterns': ['车辆信息', '基本信息', 'Vehicle Info', '基本信息表'],
                'field_mappings': {
                    'VIN': ['VIN码', '车辆识别码', 'VIN', '车架号'],
                    'make': ['品牌', '制造商', 'Make', '厂牌'],
                    'model': ['车型', '型号', 'Model', '车辆型号'],
                    'year': ['年份', 'Year', '生产年份', '年款'],
                    'production_date': ['生产日期', 'Production Date', '制造日期']
                }
            },
            'engine_info': {
                'sheet_patterns': ['发动机信息', 'Engine Info', '发动机参数'],
                'field_mappings': {
                    'engine_code': ['发动机型号', 'Engine Code', '发动机代码'],
                    'displacement': ['排量', 'Displacement', '排气量'],
                    'power': ['功率', 'Power', '额定功率'],
                    'torque': ['扭矩', 'Torque', '最大扭矩'],
                    'fuel_type': ['燃料类型', 'Fuel Type', '燃油类型']
                }
            },
            'emission_info': {
                'sheet_patterns': ['排放信息', 'Emission Info', '排放参数'],
                'field_mappings': {
                    'emission_standard': ['排放标准', 'Emission Standard', '环保标准'],
                    'co2_emission': ['CO2排放', 'CO2 Emission', '二氧化碳排放'],
                    'fuel_consumption': ['油耗', 'Fuel Consumption', '燃油消耗量']
                }
            }
        }

    def parse_file(self, file_path: str) -> Dict[str, Any]:
        """
        解析Excel文件

        Args:
            file_path: Excel文件路径

        Returns:
            解析后的数据字典
        """
        try:
            logger.info(f"开始解析Excel文件: {file_path}")

            # 获取文件基本信息
            file_info = self._get_file_info(file_path)

            # 读取所有工作表
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames

            parsed_data = {
                'file_info': file_info,
                'sheets': {},
                'structured_data': {}
            }

            # 解析每个工作表
            for sheet_name in sheet_names:
                try:
                    sheet_data = self._parse_sheet(file_path, sheet_name)
                    parsed_data['sheets'][sheet_name] = sheet_data

                    # 尝试结构化数据
                    structured_data = self._structure_sheet_data(sheet_name, sheet_data)
                    if structured_data:
                        parsed_data['structured_data'].update(structured_data)

                except Exception as e:
                    logger.error(f"解析工作表 {sheet_name} 时出错: {e}")
                    continue

            workbook.close()
            logger.info(f"Excel文件解析完成: {file_path}")
            return parsed_data

        except Exception as e:
            logger.error(f"解析Excel文件失败 {file_path}: {e}")
            raise

    def _get_file_info(self, file_path: str) -> Dict[str, Any]:
        """获取文件基本信息"""
        stat = os.stat(file_path)
        return {
            'file_name': os.path.basename(file_path),
            'file_path': file_path,
            'file_size': stat.st_size,
            'modified_time': datetime.fromtimestamp(stat.st_mtime).isoformat(),
            'file_type': 'excel'
        }

    def _parse_sheet(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """解析单个工作表"""
        try:
            # 使用pandas读取数据
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

            # 检测表头位置
            header_row = self._detect_header_row(df)

            if header_row is not None:
                # 重新读取，使用检测到的表头
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                return {
                    'type': 'structured',
                    'data': df.to_dict('records'),
                    'columns': df.columns.tolist(),
                    'shape': df.shape
                }
            else:
                # 无表头，作为原始数据
                return {
                    'type': 'raw',
                    'data': df.values.tolist(),
                    'shape': df.shape
                }

        except Exception as e:
            logger.error(f"解析工作表 {sheet_name} 失败: {e}")
            return {
                'type': 'error',
                'error': str(e)
            }

    def _detect_header_row(self, df: pd.DataFrame, max_check_rows: int = 10) -> Optional[int]:
        """检测表头位置"""
        for row_idx in range(min(max_check_rows, len(df))):
            row = df.iloc[row_idx]
            # 检查该行是否包含中文字符或常见的表头词汇
            if self._is_likely_header(row):
                return row_idx
        return 0  # 默认第一行为表头

    def _is_likely_header(self, row: pd.Series) -> bool:
        """判断是否可能是表头行"""
        header_keywords = [
            'VIN', '码', '型号', '类型', '标准', '排量', '功率', '扭矩',
            '日期', '年份', '品牌', '制造商', 'ID', 'Code', 'Name', 'Type'
        ]

        non_null_count = row.notna().sum()
        if non_null_count < len(row) * 0.3:  # 非空值少于30%
            return False

        for value in row:
            if pd.notna(value):
                value_str = str(value)
                if any(keyword in value_str for keyword in header_keywords):
                    return True
        return False

    def _structure_sheet_data(self, sheet_name: str, sheet_data: Dict[str, Any]) -> Dict[str, Any]:
        """结构化工作表数据"""
        if sheet_data['type'] != 'structured':
            return {}

        structured = {}

        # 根据配置规则匹配数据类型
        for data_type, config in self.config_rules.items():
            if self._sheet_matches_type(sheet_name, config['sheet_patterns']):
                structured_data = self._extract_structured_data(sheet_data, config)
                if structured_data:
                    structured[data_type] = structured_data
                    break

        return structured

    def _sheet_matches_type(self, sheet_name: str, patterns: List[str]) -> bool:
        """检查工作表是否匹配特定类型"""
        sheet_name_lower = sheet_name.lower()
        return any(pattern.lower() in sheet_name_lower for pattern in patterns)

    def _extract_structured_data(self, sheet_data: Dict[str, Any], config: Dict[str, Any]) -> List[Dict[str, Any]]:
        """根据配置提取结构化数据"""
        data = sheet_data['data']
        if not data:
            return []

        field_mappings = config['field_mappings']
        structured_records = []

        for record in data:
            structured_record = {}

            for target_field, source_fields in field_mappings.items():
                value = None
                for source_field in source_fields:
                    if source_field in record and pd.notna(record[source_field]):
                        value = record[source_field]
                        break

                if value is not None:
                    structured_record[target_field] = value

            if structured_record:  # 只添加非空记录
                structured_records.append(structured_record)

        return structured_records

    def configure_rules(self, rules_config: Dict[str, Any]):
        """配置解析规则"""
        self.config_rules.update(rules_config)
        logger.info("解析规则已更新")

    def save_config(self, config_path: str):
        """保存配置到文件"""
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config_rules, f, ensure_ascii=False, indent=2)
            logger.info(f"配置已保存到: {config_path}")
        except Exception as e:
            logger.error(f"保存配置失败: {e}")

    def load_config(self, config_path: str):
        """从文件加载配置"""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                self.config_rules.update(json.load(f))
            logger.info(f"配置已从文件加载: {config_path}")
        except Exception as e:
            logger.error(f"加载配置失败: {e}")

    def preview_file(self, file_path: str, max_rows: int = 10) -> Dict[str, Any]:
        """预览Excel文件内容"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames

            preview_data = {
                'file_info': self._get_file_info(file_path),
                'sheets': {}
            }

            for sheet_name in sheet_names[:3]:  # 只预览前3个工作表
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=max_rows)
                    preview_data['sheets'][sheet_name] = {
                        'columns': df.columns.tolist(),
                        'data': df.head(max_rows).fillna('').to_dict('records'),
                        'shape': df.shape
                    }
                except Exception as e:
                    preview_data['sheets'][sheet_name] = {
                        'error': str(e)
                    }

            workbook.close()
            return preview_data

        except Exception as e:
            logger.error(f"预览文件失败: {e}")
            return {'error': str(e)}

    def validate_data(self, parsed_data: Dict[str, Any]) -> Dict[str, Any]:
        """验证解析后的数据"""
        validation_result = {
            'is_valid': True,
            'errors': [],
            'warnings': [],
            'statistics': {}
        }

        try:
            # 基本验证
            if not parsed_data.get('structured_data'):
                validation_result['warnings'].append("未找到结构化数据")

            # VIN码验证
            vehicle_data = parsed_data.get('structured_data', {}).get('vehicle_info', [])
            for record in vehicle_data:
                vin = record.get('VIN', '')
                if vin and len(vin) != 17:
                    validation_result['errors'].append(f"VIN码长度不正确: {vin}")

            # 统计信息
            validation_result['statistics'] = {
                'total_sheets': len(parsed_data.get('sheets', {})),
                'structured_types': list(parsed_data.get('structured_data', {}).keys()),
                'total_records': sum(len(data) for data in parsed_data.get('structured_data', {}).values())
            }

            if validation_result['errors']:
                validation_result['is_valid'] = False

        except Exception as e:
            validation_result['is_valid'] = False
            validation_result['errors'].append(f"验证过程出错: {e}")

        return validation_result